import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from streamlit_option_menu import option_menu

path = "rapport audit.xlsx"


book = load_workbook(path)
activeSheet = book.worksheets[7]
df=pd.read_excel(path,sheet_name="FT000")
fiches=list(filter(lambda name:name.startswith("FT"),load_workbook(path).sheetnames))
programmes= pd.read_excel(path,sheet_name="Programme de travail",skiprows=7)['REF T.O'].to_list()



st.header(activeSheet['A1'].value)

def onClick():
        name=''
        for index, programme in enumerate(programmes):
                if index+1 > 9:
                    name=f"FT0{index+1}"
                else:
                    name=f"FT00{index+1}"
                sheet=book.copy_worksheet(activeSheet)
                sheet["B8"].value=programme
                sheet.title=name
                book.save(path)  
        st.success(f"Les {index} fiches sont ajoutées avec succées")

st.button("Generer les fiche de test",on_click=onClick)

def onChange():
    selected = option_menu(
        menu_title=None,
        options=["Saisie des données", ],
        orientation="horizontal")
    if option_menu == "Afficahge des données":
        df=pd.read_excel(path,st.session_state['selected_fiche'])
        st.table(df.dropna(subset = ['ANNEXE N°06 : FICHE DE TEST']))
    else:
        st.form("")

  
st.selectbox("Fiches de test",key="selected_fiche", options=fiches,on_change=onChange,placeholder="Choisir")
